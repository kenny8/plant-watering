"""
Утилиты для экспорта данных в формат Excel (.xlsx).
Использует openpyxl для генерации файлов в буфере io.BytesIO.
"""
import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from bot.core.database import Database
from bot.utils.logger import setup_logger

logger = setup_logger(__name__)


def generate_excel_buffer(
    session: object,
    device_id: int,
    build_id: int,
    field_name: str
) -> io.BytesIO:
    """
    Генерирует Excel-файл с показаниями датчика в буфере BytesIO.
    
    Args:
        session: SQLAlchemy session или connection
        device_id: ID устройства
        build_id: ID сборки
        field_name: Имя поля (датчика)
    
    Returns:
        io.BytesIO: Буфер с Excel-файлом
    
    Структура файла:
        - Заголовок: "📊 Показания {field_name}"
        - Колонки: "Время записи", "Значение"
        - Формат даты: DD.MM.YYYY HH:MM:SS
        - Автоширина колонок
    """
    logger.info(f"[EXPORT] Начало генерации Excel: device_id={device_id}, build_id={build_id}, field_name='{field_name}'")
    
    # Создаем буфер в памяти
    buffer = io.BytesIO()
    
    # Создаем книгу и лист
    wb = Workbook()
    ws = wb.active
    ws.title = "Показания"[:31]  # Лимит имени листа в Excel
    
    # Заголовок документа (без иконки)
    field_display = " ".join(word.capitalize() for word in field_name.replace('_', ' ').split())
    ws.merge_cells('A1:B1')
    ws['A1'] = f"Показания: {field_display}"
    ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
    ws['A1'].alignment = ws['A1'].alignment.copy(horizontal='center')
    
    # Заголовки колонок
    ws['A2'] = "Время записи"
    ws['B2'] = "Значение"
    
    # Стиль заголовков колонок
    for cell in ['A2', 'B2']:
        ws[cell].font = ws[cell].font.copy(bold=True)
        ws[cell].alignment = ws[cell].alignment.copy(horizontal='center')
    
    # Получаем данные из БД
    readings = []
    try:
        # Определяем тип сессии (connection или session)
        if hasattr(session, 'execute'):
            conn = session
        else:
            conn = session
        
        result = conn.execute(
            text("""
                SELECT created_at, field_value
                FROM device_data
                WHERE device_id = :device_id 
                  AND build_id = :build_id 
                  AND field_name = :field_name
                ORDER BY created_at ASC
            """),
            {
                "device_id": device_id,
                "build_id": build_id,
                "field_name": field_name
            }
        )
        rows = result.fetchall()
        readings = [(row[0], row[1]) for row in rows]
        logger.info(f"[EXPORT] Получено {len(readings)} записей из БД")
        
    except Exception as e:
        logger.error(f"[EXPORT] Ошибка выполнения SQL-запроса: {e}")
        raise
    
    # Заполняем данными
    if not readings:
        logger.warning(f"[EXPORT] Нет данных для экспорта")
        ws.merge_cells('A3:B3')
        ws['A3'] = "📭 Нет данных для экспорта"
        ws['A3'].alignment = ws['A3'].alignment.copy(horizontal='center')
        ws['A3'].font = ws['A3'].font.copy(italic=True, color="FF666666")
    else:
        logger.debug(f"[EXPORT] Запись {len(readings)} строк в Excel")
        for idx, (created_at, field_value) in enumerate(readings, start=3):
            # Форматируем дату
            if isinstance(created_at, datetime):
                date_str = created_at.strftime("%d.%m.%Y %H:%M:%S")
            elif created_at:
                date_str = str(created_at)[:19]
            else:
                date_str = "N/A"
            
            # Значение
            value_str = str(field_value) if field_value is not None else "N/A"
            
            ws[f'A{idx}'] = date_str
            ws[f'B{idx}'] = value_str
        
        # Применяем автоширину колонок
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 25
    
    # Сохраняем в буфер
    wb.save(buffer)
    buffer.seek(0)
    
    file_size = buffer.tell()
    logger.info(f"[EXPORT] Excel-файл сгенерирован успешно (размер: {file_size} байт)")
    
    return buffer


def format_filename(field_name: str, device_id: int) -> str:
    """
    Формирует имя файла для экспорта.
    
    Args:
        field_name: Имя поля
        device_id: ID устройства
    
    Returns:
        str: Имя файла в формате "sensor_{field_name}_{device_id}.xlsx"
    """
    # Очищаем имя поля от недопустимых символов
    safe_name = field_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
    safe_name = "".join(c for c in safe_name if c.isalnum() or c in "_-")[:30]
    
    filename = f"sensor_{safe_name}_{device_id}.xlsx"
    logger.debug(f"[EXPORT] Имя файла: {filename}")
    
    return filename
